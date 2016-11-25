from __future__ import print_function
from openpyxl import Workbook, load_workbook
import xlrd
import json
import csv
import warnings

# CALL THIS ONLY
# @param: spreadsheet => the full path to a .xls .xlsx or .csv file
#
def inputSpreadSheet(spreadsheet):
  data = None
  if spreadsheet.endswith(".xls"):
    data = xlsParser(spreadsheet)
  elif spreadsheet.endswith(".xlsx"):
    with warnings.catch_warnings():
      warnings.simplefilter("ignore")
      data = xlsxParser(spreadsheet)
  elif spreadsheet.endswith(".csv"):
    data = csvParser(spreadsheet)
  else:
    pass
    # TODO: raise exception or return an error

  validate(data) # have this throw exception

  json_arr = json.dumps(data)

  json_data = '{ "Columns" : ["Product", "Location", "Quantity", "Price"], "Data" : ' + json_arr + ' }'

  return json_data

# EVERYTHING BEYOND THIS FACILITATES PARSING, AND SHOULD NOT BE CALLED OUTSITE THIS FILE

def xlsxParser(xl_file):
  wb = load_workbook(filename=xl_file) #can add read_only=True
  names = wb.get_sheet_names()
  ws = wb[names[0]]
  data = []

  for row in ws.iter_rows(max_col=4):
    record = []
    for item in row:
      cell = item.value
      if type(cell) is unicode:
        cell = str(cell)
        if cell == "":
          cell = None
      if type(cell) is str:
        cell = cell.strip()
      record.append(cell)
    if isRowValid(record): data.append(record)

  return data


def xlsParser(xl_file):
  wb = xlrd.open_workbook(xl_file)
  #xlsPrintBook(wb)

  ws = wb.sheet_by_index(0)
  data = []
  rowCount = ws.nrows

  for i_row in range(0,rowCount):
    record = []

    for i_col in range(0,4):
      cell = (ws.cell(i_row, i_col)).value
      if type(cell) is unicode:
        cell = str(cell)
        if cell == "":
          cell = None
      if type(cell) is str:
        cell = cell.strip()
      record.append(cell)

    if isRowValid(record):data.append(record)

  return data


def csvParser(csv_file):
  data = []
  with open(csv_file, "rU") as csv_f:
    csv_read = csv.reader(csv_f)

    for row in csv_read:
      record = []

      for item in row:
        if item == "":
          item = None
        elif item.startswith("$"):
          item = float(item[1:])
        if type(item) is str:
          item = item.strip()
        record.append(item)

      if isRowValid(record):data.append(record)

  return data


# UTILITY FUNCTIONS
def isRowValid(row):  
  if row[0] == None:
    return False

  if row[3] == None or type(row[3]) == str:
    return False

  return True


def validate(data):
  pass 
  # TODO: implement validation


# DEBUGGING FUNCTIONS
def xlsxPrintBook(wb):
  ws_names = wb.get_sheet_names()
  for name in ws_names:
    ws = wb[name]
    print("For Work Sheet: {}".format(name))
    for row in ws.iter_rows():
      for cell in row:
        print("[{}]".format(cell.value), end="")
      print(" ", end="\n")


def xlsPrintBook(wb):
  ws_names = wb.sheet_names()
  for name in ws_names:
    ws = wb.sheet_by_name(name)
    print("For Work Sheet: {}".format(name))
    colCount = ws.ncols
    rowCount = ws.nrows
    for row_idx in range(0, rowCount):
      for col_idx in range(0, colCount):
        cell = ws.cell(row_idx, col_idx)
        print("[{}]".format(cell), end="")
      print(" ", end="\n")


